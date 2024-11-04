import sys, random
from PyQt5.QtWidgets import QMainWindow,QApplication,QWidget
from openpyxl import load_workbook
from Ui_Form import Ui_Form

class ExcelData():
    def change_row():
        workbook = load_workbook(filename='.\Data.xlsx')
        sheet = workbook['Sheet1']
        non_empty_rows = [i for i in range(1, sheet.max_row + 1) if sheet[f'A{i}'].value]
        selected_row_num = random.choice(non_empty_rows)

        return selected_row_num
        
    def get(cell_name):
        # 加载工作簿
        workbook = load_workbook(filename='.\Data.xlsx')
        # 选择Sheet1
        sheet = workbook['Sheet1']
        # 获取指定单元格的内容
        cell_content = sheet[cell_name].value
        # 返回单元格内容
        return cell_content
    

class MyMainWindow(QMainWindow,Ui_Form):
    def __init__(self,card,parent =None):
        self.card = card
        super(MyMainWindow,self).__init__(parent)
        self.setupUi(self)
        self.btn_next.clicked.connect(self.next)
        self.btn_answer.clicked.connect(self.answer)

    def next(self):
        if (self.txt_next.toPlainText() == ''):
            self.txt_next.setText(ExcelData.get('A' + self.card))
        else:
            self.card = (str)(ExcelData.change_row())
            self.txt_next.setText(ExcelData.get('A' + self.card))

    def answer(self):
        self.txt_answer.setText(ExcelData.get('B' + self.card))


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWin = MyMainWindow('1')
    myWin.show()
    sys.exit(app.exec_())    
